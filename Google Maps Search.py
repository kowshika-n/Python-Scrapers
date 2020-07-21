# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
import unittest
import openpyxl
import time
import os
import re

home = os.environ.get('USERPROFILE')
ChromePath = '--user-data-dir=' + home + '\\AppData\\Local\\Google\\Chrome\\User Data\\'
xl_path = '.\Places.xlsx'

Place1 = 'New York, USA'
Place2 = 'California, USA'

# Load the workbook from path
# Dont use_iterators=True when read_only=False. File Wont save!
wb = openpyxl.load_workbook(xl_path, read_only=False, data_only=True)
# Open the worksheet from the name
sheet = wb.get_sheet_by_name('Input')

# Get Row and Column count
row_count = sheet.max_row
column_count = sheet.max_column
colNum = column_count + 1

print('\n\tNo. of rows = %r\n\t columns = %r' % (row_count, column_count))
print('Total no. of combinations = ', (row_count * column_count))

# Each row in the spreadsheet has data and is appended to the list
Places1 = []
Places2 = []

# skip the first row and empty rows with for if loop.
Places1 = [str(sheet['A' + str(row)].value) for row in range(2,row_count + 1) if sheet['A' + str(row)].value is not None]
for column in range(1, 2):
    for cell in sheet[str(column)]:
        if cell.value is not None:
            Places2.append(cell.value)
wb.save(xl_path)


class GMapsSearch(unittest.TestCase):
    def setUp(self):
        chromedriver = home + "\chromedriver.exe"
        options = webdriver.ChromeOptions()
        # Path to your chrome profile
        options.add_argument(ChromePath)

        self.driver = webdriver.Chrome()
        self.driver.implicitly_wait(30)
        self.base_url = "https://www.google.co.in/"
        self.verificationErrors = []
        self.accept_next_alert = True

    def test_g_maps_search(self):
        driver = self.driver
        driver.get(self.base_url + "/maps/@19.1598727,72.9992012,15z?hl=en")
        driver.find_element_by_id("searchbox-directions").click()
        time.sleep(2)
        for i in range(60):
            try:
                if self.is_element_present(By.XPATH, "//div[@id='omnibox-directions']/div/div[2]/div/div/div/div[2]/button"): break
            except: pass
            time.sleep(1)
        else: self.fail("time out")

        time.sleep(2)
        driver.find_element_by_xpath("//div[@id='omnibox-directions']/div/div[2]/div/div/div/div[2]/button").click()

        for Place1 in Places1:
            for Place2 in Places2:
                driver.find_element_by_css_selector("#sb_ifc51 > input.tactile-searchbox-input").clear()
                driver.find_element_by_css_selector("#sb_ifc51 > input.tactile-searchbox-input").send_keys(Place1)
                driver.find_element_by_css_selector("#sb_ifc52 > input.tactile-searchbox-input").clear()
                driver.find_element_by_css_selector("#sb_ifc52 > input.tactile-searchbox-input").send_keys(Place2)
                time.sleep(0.5)
                driver.find_element_by_css_selector("#directions-searchbox-1 > button.searchbox-searchbutton").click()
                time.sleep(0.5)
                for i in range(60):
                    try:
                        if self.is_element_present(By.CSS_SELECTOR, "div.section-directions-trip-distance.section-directions-trip-secondary-text > div"): break
                    except: pass
                    time.sleep(1)
                else: self.fail("time out")
                Dist = driver.find_element_by_css_selector("div.section-directions-trip-distance.section-directions-trip-secondary-text > div").text

                if 'km' in Dist:
                    Dist = Dist.replace('km', '')
                if ',' in Dist:
                    Dist = Dist.replace(',', '')
                print(Place1 + ' , ' + Place2 + ' , ' + Dist)

                driver.find_element_by_css_selector("#sb_ifc51 > input.tactile-searchbox-input").clear()
                time.sleep(1)


    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException as e: return False
        return True

    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException as e: return False
        return True

    def close_alert_and_get_its_text(self):
        try:
            alert = self.driver.switch_to_alert()
            alert_text = alert.text
            if self.accept_next_alert:
                alert.accept()
            else:
                alert.dismiss()
            return alert_text
        finally: self.accept_next_alert = True

    def tearDown(self):
        self.driver.quit()
        self.assertEqual([], self.verificationErrors)


if __name__ == "__main__":
    unittest.main()
