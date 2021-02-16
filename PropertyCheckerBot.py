from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
import openpyxl
import os
import time


xl_path = "Input.xlsx"
PIDsiteURL = ADRsiteURL = ""

os.environ['WDM_LOG_LEVEL'] = '0'
options = webdriver.ChromeOptions()
options.add_argument("--disable-notifications")
options.add_argument("--start-maximized")
options.add_argument("--disable-popups")
DEFAULT_IMPLICIT_WAIT = 10


def GetElement(driver, elementTag, locator='ID'):
    '''Wait for element and then return when it is available'''
    try:
        if locator == 'ID':
            if is_element_present(driver, By.ID, elementTag):
                return WebDriverWait(driver, 15).until(
                    lambda driver: driver.find_element_by_id(elementTag))
            else:
                print('%s Not Found.' % elementTag)
                return None

        elif locator == 'NAME':
            if is_element_present(driver, By.NAME, elementTag):
                return WebDriverWait(driver, 15).until(
                    lambda driver: driver.find_element_by_name(elementTag))
            else:
                print('%s Not Found.' % elementTag)
                return None

        elif locator == 'XPATH':
            if is_element_present(driver, By.XPATH, elementTag):
                return WebDriverWait(driver, 15).until(
                    lambda driver: driver.find_element_by_xpath(elementTag))
            else:
                print('%s Not Found.' % elementTag)
                return None

        elif locator == 'CSS':
            if is_element_present(driver, By.CSS_SELECTOR, elementTag):
                return WebDriverWait(driver, 15).until(
                    lambda driver: driver.find_element_by_css_selector(elementTag))
            else:
                print('%s Not Found.' % elementTag)
                return None

    except Exception as e:
        print('Error identifying element - %s.' % e)
    return None


def is_element_present(driver, how, what):
    """Check if an element is present"""
    try:
        driver.find_element(by=how, value=what)
    except NoSuchElementException:
        return False
    return True


def WaitTillElementPresent(driver, elementTag, locator='ID'):
    '''Wait till element present. Max 30 seconds'''
    result = False
    driver.implicitly_wait(0)
    for i in range(30):
        try:
            if locator == 'ID':
                if is_element_present(driver, By.ID, elementTag):
                    result = True
                    break
            elif locator == 'NAME':
                if is_element_present(driver, By.NAME, elementTag):
                    result = True
                    break
            elif locator == 'XPATH':
                if is_element_present(driver, By.XPATH, elementTag):
                    result = True
                    break
            elif locator == 'CSS':
                if is_element_present(driver, By.CSS_SELECTORS, elementTag):
                    result = True
                    break
        except Exception as e:
            print('Exception when WaitTillElementPresent : %s' % e)
            pass
        time.sleep(0.99)
    else:
        print("Timed out. Unable to find the Element: %s" % elementTag)
    driver.implicitly_wait(DEFAULT_IMPLICIT_WAIT)
    return result


def Type(driver, elementTag, locator='ID', input_text=''):
    """Type the given input text"""
    try:
        WaitTillElementPresent(driver, elementTag, locator)
        element = GetElement(driver, elementTag, locator)
        element.click()
        element.clear()
        element.send_keys(input_text)
    except Exception as e:
        print('Exception when Type : %s' % e)
        pass


def Click(driver, elementTag, locator='ID'):
    """Click any given element"""
    try:
        WaitTillElementPresent(driver, elementTag, locator)
        element = GetElement(driver, elementTag, locator)
        element.click()
    except Exception as e:
        print('Exception when Click : %s' % e)
        pass


def scrollTo(driver, elementTag, locator='ID'):
    """Scroll to any given element"""
    try:
        element = GetElement(driver, elementTag, locator)
        if element:
            driver.execute_script("arguments[0].scrollIntoView({block: \"center\", inline: \"center\"});", element)
    except Exception as e:
        print('Exception when scrolling : %s' % e)
        pass


def tearDown(driver):
    """Close the browser and driver at the end of script"""
    try:
        driver.close()
    except Exception as e:
        pass
    try:
        driver.quit()
    except Exception as e:
        pass


def is_alert_present(driver):
    """Function to find if any alert is visible"""
    try:
        driver.switch_to_alert
    except NoAlertPresentException:
        return False
    return True


def close_alert(driver):
    """Function to close alert if any"""
    alert_text = ''
    try:
        if (is_alert_present(driver)):
            alert = driver.switch_to.alert
            alert_text = alert.text
            print("Alert Message : " + alert_text)
            alert.dismiss()
        if (is_alert_present(driver)):
            alert = driver.switch_to.alert
            print("Alert Message : " + alert.text)
            alert.accept()
        return alert_text
    except Exception as e:
        #print('Exception when Alert : %s' % e)
        pass


def Found(filename):
    """
    Function to find if an input file exists.
    Give a file path and output will be True or False.
    """
    try:
        if filename is None:
            print("Input Filename is None.")
            return False
        else:
            if os.path.exists(filename) is True:
                if os.path.isfile(filename):
                    # print('%s is available.' % os.path.basename(filename))
                    return True
                else:
                    print('Error:Input Path %s is not a file but a folder.' % filename)
                    return False
            else:
                print('Error: %s Doesnot Exist. Or Path is incorrect.' % filename)
                return False
    except:
        print('Error: %s Doesnot Exist. Or Path is incorrect.' % filename)
        return False


def Closed(filename):
    """
    Function to check if an input file exists and not open.
    Returns True if the File is closed
    Returns False if the File is Open/Not Found.
    """
    if Found(filename) is True:
        try:
            os.rename(filename, filename)
            # print('%s is Closed!' % (os.path.basename(filename)))
            return True
        except Exception as e:
            print('Error Accessing the file! File may be Open!\n' + str(e))
    return False


def main():
    if not Closed(xl_path):
        print('Please close the file and re-run this bot.')
    else:
        wb = None
        try:
            driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), options=options)
            driver.maximize_window()
            print("Google Chrome Launched!")

            # Load the workbook from path
            # Dont use_iterators=True when read_only=False. File Wont save!
            wb = openpyxl.load_workbook(xl_path, read_only=False, data_only=True)
            # Open the worksheet from the name
            print(f"Excel {xl_path} opened. Sheets found : {wb.sheetnames}")

            #Take the first sheet name
            sheet = wb[wb.sheetnames[0]]

            # Get Row and Column count
            row_count = sheet.max_row
            column_count = sheet.max_column
            colNum = column_count + 1
            print('No. of rows = %r and columns = %r' % (row_count, column_count))

            for row in range(2, row_count + 1):
                PID = sheet['C' + str(row)].value
                if PID and len(PID) >= 13:
                    siteURL = ADRsiteURL
                    if "-" in PID:
                        p = PID.replace("-", "").strip()
                        if p.isnumeric():
                            siteURL = PIDsiteURL
                    
                    driver.get(siteURL)
                    driver.implicitly_wait(DEFAULT_IMPLICIT_WAIT)
                    WaitTillElementPresent(driver, "//input", locator='XPATH')

                    if siteURL == PIDsiteURL:
                        Type(driver, "//*[@name='pid']", locator='XPATH', input_text=PID)
                        Click(driver, "//*[@value='Search']", locator='XPATH')
                        close_alert(driver)

                    if siteURL == ADRsiteURL:
                        AdressList = PID.strip().split(" ")
                        HOUSE = ADDRESS = UNIT = ""
                        if AdressList[0].strip().isnumeric():
                            HOUSE = AdressList[0].strip()
                            if "#" in AdressList[-1]:
                                UNIT = AdressList[-1].replace("#", "").strip()
                                del AdressList[-1]
                            ADDRESS = " ".join(AdressList[1:])
                        else:
                            HOUSE = "1"
                            if "#" in AdressList[-1]:
                                UNIT = AdressList[-1].replace("#", "").strip()
                                del AdressList[-1]
                            ADDRESS = " ".join(AdressList)
                        if ADDRESS:
                            if HOUSE:
                                Type(driver, "//input[@name='house']", locator='XPATH', input_text=HOUSE)
                            Type(driver, "//input[@name='street']", locator='XPATH', input_text=ADDRESS)
                            if UNIT:
                                Type(driver, "//input[@name='condo']", locator='XPATH', input_text=UNIT)
                            Click(driver, "//*[@value='Search']", locator='XPATH')
                            close_alert(driver)

                    WaitTillElementPresent(driver, "//*[contains(text(), 'Owner name')] | //*[text() = 'No records found']", locator='XPATH')


        except Exception as e:
            print(str(e))
        finally:
            tearDown(driver)
            # Save the workbook.
            wb.save(xl_path)


if __name__ == '__main__':
    main()

