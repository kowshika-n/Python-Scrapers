from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime
from glob import glob
import configparser
import random
import openpyxl
import sys
import re
import os
import time
import json

current_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(current_dir)
siteURL  = "https://www.outdoorsy.com/"
xl_path  = "Output.xlsx"
_author  = "Kowshika @ https://kowshika-n.github.io/"
_version = 0.1

os.environ['WDM_LOG_LEVEL'] = '0'
options = webdriver.ChromeOptions()
options.add_argument("--disable-notifications")
options.add_argument("--start-maximized")       # ("--kiosk") for MAC
options.add_argument("--disable-popups")
options.add_argument("user-agent=Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)")
options.add_argument('--no-sandbox')   # Bypass OS security model
#options.add_argument("--headless")     # Runs Chrome in headless mode.
#options.add_argument('--disable-gpu')  # applicable to windows os only

DEFAULT_IMPLICIT_WAIT = 30


def getDateTime():
    """Returns timestamp in string format"""
    return datetime.today().strftime('%m%d%Y_%H%M%S')


def catch(error):
    '''Method to catch errors and display error line'''
    exc_type, exc_obj, exc_tb = sys.exc_info()
    lineNo = str(exc_tb.tb_lineno)
    print('%s : %s at Line %s.' % (type(error), error, lineNo))


def GetElement(driver, elementTag, locator='ID'):
    '''Wait for element and then return when it is available'''
    try:
        locator = locator.upper()
        if locator == 'ID':
            if is_element_present(driver, By.ID, elementTag):
                return WebDriverWait(driver, 15).until(
                    lambda driver: driver.find_element_by_id(elementTag))
            else:
                print('%s Not Found.' % elementTag)
                return None

        elif locator == 'NAME':
            if is_element_present(driver, By.NAME, elementTag):
                return WebDriverWait(driver, 15).until(
                    lambda driver: driver.find_element_by_name(elementTag))
            else:
                print('%s Not Found.' % elementTag)
                return None

        elif locator == 'XPATH':
            if is_element_present(driver, By.XPATH, elementTag):
                return WebDriverWait(driver, 15).until(
                    lambda driver: driver.find_element_by_xpath(elementTag))
            else:
                print('%s Not Found.' % elementTag)
                return None

        elif locator == 'CSS':
            if is_element_present(driver, By.CSS_SELECTOR, elementTag):
                return WebDriverWait(driver, 15).until(
                    lambda driver: driver.find_element_by_css_selector(elementTag))
            else:
                print('%s Not Found.' % elementTag)
                return None

    except Exception as e:
        print("Element not found with %s : %s" % (locator, elementTag))
    return None



def GetElements(driver, elementTag, locator='ID'):
    '''Wait for element and then return the list when available'''
    try:
        locator = locator.upper()
        if locator == 'ID':
            if is_element_present(driver, By.ID, elementTag):
                return WebDriverWait(driver, 15).until(
                    lambda driver: driver.find_elements_by_id(elementTag))
            else:
                print('%s Not Found.' % elementTag)
                return None

        elif locator == 'NAME':
            if is_element_present(driver, By.NAME, elementTag):
                return WebDriverWait(driver, 15).until(
                    lambda driver: driver.find_elements_by_name(elementTag))
            else:
                print('%s Not Found.' % elementTag)
                return None

        elif locator == 'XPATH':
            if is_element_present(driver, By.XPATH, elementTag):
                return WebDriverWait(driver, 15).until(
                    lambda driver: driver.find_elements_by_xpath(elementTag))
            else:
                print('%s Not Found.' % elementTag)
                return None

        elif locator == 'CSS':
            if is_element_present(driver, By.CSS_SELECTOR, elementTag):
                return WebDriverWait(driver, 15).until(
                    lambda driver: driver.find_elements_by_css_selector(elementTag))
            else:
                print('%s Not Found.' % elementTag)
                return None

    except Exception as e:
        print("Elements not found with %s : %s" % (locator, elementTag))
    return None


def is_element_present(driver, how, what):
    """Check if an element is present"""
    try:
        driver.find_element(by=how, value=what)
    except NoSuchElementException:
        return False
    return True

def tearDown(driver):
    """Close the browser and driver at the end of script"""
    try:
        driver.close()
    except Exception as e:
        pass
    try:
        driver.quit()
    except Exception as e:
        pass


def selectTrailers(driver, Vehicletypes, index):
    Vehicletype = ""
    # uncheck map
    mapCheckbox = "//div[contains(@class, 'map')]/div[@role='checkbox' and @aria-checked='true']"
    if WaitTillElementPresent(driver, mapCheckbox, 'xpath', 2):
        ActionClick(driver, mapCheckbox, 'xpath')
        time.sleep(3)

    # get vehicle type and select one by one
    if index < len(Vehicletypes) and Vehicletypes[index]:
        Vehicletype = Vehicletypes[index].split(",")
    else:
        Vehicletype = Vehicletype.split(",")

    if Vehicletype and len(Vehicletype) > 0 and Vehicletype[0] != '':
        ActionClick(driver, "//div[text() = 'Vehicle type']", 'xpath')
        WaitTillElementPresent(driver, "//div[text() = 'All towables']", 'xpath')
        for Type in Vehicletype:
            VehicleXpath = "//img[@alt='" + Type + "']//ancestor::div[contains(@class,'box') and @data-test-checkbox-image]"
            ActionClick(driver, VehicleXpath, 'xpath')
            time.sleep(2)
        Click(driver, "//button//*[text()='Apply']", 'xpath')
    time.sleep(3)


def getReviews(MinNumberofReviews, index):
    Reviews = ""
    if index < len(MinNumberofReviews) and MinNumberofReviews[index]:
        Reviews = MinNumberofReviews[index]
        Reviews = int(Reviews)
    else:
        Reviews = 1
    return Reviews


def main():
    """Extraction Begins"""
    new_xl = getNewExcelName(xl_path)
    
    if not Closed(xl_path):
        print('Please close the file and re-run this bot.')
    else:
        wb = None
        try:
            driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), options=options)
            driver.maximize_window()
            print("Google Chrome Launched!")
            
            # Load the workbook from path
            # Dont use_iterators=True when read_only=False. File Wont save!
            wb = openpyxl.load_workbook(xl_path, read_only=False, data_only=True)
            # Open the worksheet from the name
            print(f"Excel {xl_path} opened. Sheets found : {wb.sheetnames}")

            #Take the first sheet name
            sheet = wb[wb.sheetnames[0]]

            # Get Row and Column count
            row_count = sheet.max_row
            column_count = sheet.max_column
            colNum = column_count + 1
            print('No. of rows = %r and columns = %r' % (row_count, column_count))

            driver.implicitly_wait(DEFAULT_IMPLICIT_WAIT)
            row = row_count

            Input_sheet = wb['Input']
            Input_row_count = Input_sheet.max_row
            Locations, Vehicletypes , MinNumberofReviews = ([] for i in range(3))
            # exclude first row and read remaining rows
            for i in range(2, Input_row_count+1):
                Locations.append(Input_sheet['A'+ str(i)].value)
                Vehicletypes.append(Input_sheet['B'+ str(i)].value)
                MinNumberofReviews.append(Input_sheet['C'+ str(i)].value)
            print(Locations, Vehicletypes, MinNumberofReviews)

            for index, location in enumerate(Locations):
                redirectFromGoogle(driver, siteURL)
                WaitTillElementPresent(driver, 'header', 'ID')
                WaitTillElementPresent(driver, "//div/input[@type='search']", 'xpath')
                ActionTypeSlow(driver, "//div/input[@type='search']", 'xpath', input_text=location)
                time.sleep(2)
                # click on first result and search
                jsClick(driver, "(//div[@data-test-mapbox-suggestion])[1]", 'xpath')
                time.sleep(1)
                ActionClick(driver, "//span[text() = 'Search']", 'xpath')
                WaitTillElementPresent(driver, "//*[text() = 'More filters' or text() = 'Map view']", 'xpath')

                selectTrailers(driver, Vehicletypes, index)
                Reviews = getReviews(MinNumberofReviews, index)
                    
                rentalcard_xpath = "//div[@data-test-rental-card]"
                WaitTillElementPresent(driver, rentalcard_xpath, 'xpath', 10)
                cards = getElementCount(driver, rentalcard_xpath, 'xpath')
                for i in range(1, cards+1):
                    card_xpath = "(" + rentalcard_xpath + ")[" + str(i) + "]" 
                    if WaitTillElementPresent(driver, card_xpath+ "//div[contains(@class, 'rating')]", 'xpath', 2):
                        review_count = getText(driver, card_xpath + "//div[contains(@class, 'rating')]//span[@id]", 'xpath')
                        review_count = int(getNumbersOnly(review_count))
                        if review_count >= Reviews:
                            vehicle_link_xpath = card_xpath + "//a"
                            WaitTillElementPresent(driver, vehicle_link_xpath, 'xpath', 5)
                            ActionClick(driver, vehicle_link_xpath, 'xpath')
                            time.sleep(5)
                            driver.switch_to.window(driver.window_handles[-1])

                            WaitTillElementPresent(driver, '//a[@type="button"]', 'xpath')
                            WaitTillElementPresent(driver, '//ul[@id]', 'xpath')
                            URL = Location = ListingTitle = NightlyFee = MinimumStay = SecurityDeposit = Availability = ""
                            AvailabileDates = Numberofreviews = AverageStars = ImageCount = Description = ""

                            URL = driver.current_url
                            Location = getText(driver, "//h1//following::div[1]", 'xpath').replace('\\n', '').replace('\n', '')
                            ListingTitle = getText(driver, "//h1[@data-test-rv-show-title]", 'xpath')
                            NightlyFee = getText(driver, "//div[contains(@class,'price')]/em", 'xpath')
                            MinimumStay = getNumbersOnly(getText(driver, "//div[contains(text(), 'night minimum') or contains(text(), 'nights minimum')]", 'xpath'))
                            SecurityDeposit = getText(driver, "(//*[contains(text(), 'deposit')]/following::text())[2]", 'xpath')
                            #ListingTitle
                            
                            print(f'{Location} {URL}')

                            driver.close()
                            driver.switch_to.window(driver.window_handles[0])
                            time.sleep(3)
                        else:
                            print(f'Reviews:{review_count} less than expected {Reviews} Reviews')

        except Exception as e:
            catch(e)
        finally:
            tearDown(driver)


if __name__ == '__main__':
    main()
    time.sleep(5)
